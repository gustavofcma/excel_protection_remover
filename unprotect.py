import PySimpleGUI as sg
from openpyxl import load_workbook

NOT_BOOLEAN = ['spinCount','hashValue','saltValue','algorithmName']
SUFIXO_PADRAO = '_unlock'


def get_file():
    form_rows = [[sg.Text('Selecione o arquivo desejado')],
                 [sg.Text('Selecionar'),
                    sg.InputText(key='-arquivo-'), sg.FileBrowse('Buscar', file_types=(("Arquivos Excel", "*.xlsx"),))],
                 [sg.Text('Sufixo'), sg.InputText(key='-sufixo-', default_text=SUFIXO_PADRAO, size=(15, 1))],
                 [sg.Submit('Desproteger'), sg.Cancel('Cancelar')]]

    window = sg.Window('Desproteger arquivo Excel', form_rows)
    event, values = window.read()
    window.close()
    return event, values


def unprotect(file_path, suffix):
	new_file_path = f'{suffix}.'.join(file_path.rsplit('.', 1))

	workbook = load_workbook(file_path)
	worksheet_names = workbook.sheetnames

	for worksheet_name in worksheet_names:

		worksheet = workbook[worksheet_name]
		worksheet.sheet_state = 'visible'

		for item in worksheet.protection:
			
			if item[0] in NOT_BOOLEAN:
				worksheet.protection.__setattr__(item[0], None)
			else:
				worksheet.protection.__setattr__(item[0], False)

	workbook.save(new_file_path)

	return new_file_path


def main():
	button, values = get_file()
	excel_file_path = values['-arquivo-']
	sufixo = values['-sufixo-']

	if any((button != 'Desproteger', excel_file_path == '', sufixo == '')):
		sg.popup_error('Operação cancelada')
		return

	print(excel_file_path)

	new_file = unprotect(excel_file_path, sufixo)

	sg.popup(f'Novo arquivo criado.\nCaminho: {new_file}')
	#while True:
	#	event, values = window.read()
	#	if event == sg.WIN_CLOSED or event == 'Cancel': # if user closes window or clicks cancel
	#		break


if __name__ == '__main__':
	main()
